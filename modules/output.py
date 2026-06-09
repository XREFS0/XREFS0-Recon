import json
import csv
import os
from datetime import datetime
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.layout import Layout
from rich.text import Text
from rich import box

console = Console()

class OutputEngine:
    def __init__(self, data, findings=None):
        self.data = data
        self.domain = data.get("domain", "unknown")
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self._findings = findings or []

    def print_summary(self):
        console.print(Panel(f"[bold red]XREFS0 - Cyber Intelligence & Domain Recon Platform[/bold red]",
                           subtitle=f"[bold yellow]Target: {self.domain}[/bold yellow] | {self.timestamp}",
                           border_style="red",
                           box=box.DOUBLE_EDGE))

    def print_subdomains_table(self, subdomains, dns_results=None, http_results=None):
        if not subdomains:
            console.print("[red]No subdomains found.[/red]")
            return
        table = Table(title=f"Subdomains ({len(subdomains)} found)", box=box.ROUNDED, header_style="bold cyan")
        table.add_column("Subdomain", style="white")
        table.add_column("IP", style="green")
        table.add_column("Status", style="yellow")
        table.add_column("CDN", style="magenta")
        table.add_column("Server", style="blue")
        table.add_column("Title", style="cyan")
        for sub in subdomains[:100]:
            ip = ""
            status = ""
            cdn = ""
            server = ""
            title = ""
            if dns_results and sub in dns_results:
                dns = dns_results[sub]
                ips = dns.get("A", [])
                if ips:
                    ip = ips[0]
                elif dns.get("AAAA"):
                    ip = dns["AAAA"][0]
                else:
                    ip = "-"
            if http_results and sub in http_results:
                h = http_results[sub]
                status = str(h.get("status_code", ""))
                server = h.get("server", "")
                title = h.get("title", "")
                if h.get("alive"):
                    status = f"[green]{status}[/green]"
                else:
                    status = f"[red]{status}[/red]"
                cdn_data = h.get("cdn", [])
                if cdn_data:
                    cdn = cdn_data[0].get("name", "") if isinstance(cdn_data, list) and len(cdn_data) > 0 else ""
            table.add_row(sub, ip, status, cdn, server[:30], title[:40] if title else "")
        console.print(table)

    def print_whois(self, whois_data):
        if not whois_data:
            return
        if not whois_data.get('is_registered') and not whois_data.get('registrar'):
            status_text = "[yellow]WHOIS lookup failed or domain not found[/yellow]"
        elif whois_data.get('is_registered') and not whois_data.get('is_expired'):
            status_text = "[green]ACTIVE[/green]"
        elif whois_data.get('is_expired'):
            status_text = "[red]EXPIRED[/red]"
        else:
            status_text = "[yellow]Unknown[/yellow]"
        panel = Panel(
            f"[bold]Registrar:[/bold] {whois_data.get('registrar') or 'N/A'}\n"
            f"[bold]Created:[/bold] {whois_data.get('created') or 'N/A'}\n"
            f"[bold]Expires:[/bold] {whois_data.get('expires') or 'N/A'}\n"
            f"[bold]Status:[/bold] {status_text}\n"
            f"[bold]Days until expiry:[/bold] {whois_data.get('days_until_expiry') or 'N/A'}",
            title="WHOIS Information",
            box=box.HEAVY,
            border_style="yellow",
        )
        console.print(panel)

    def print_real_ip(self, real_ip_data):
        if not real_ip_data:
            return
        ips = real_ip_data.get("ips", [])
        methods = real_ip_data.get("methods", {})
        if ips:
            panel = Panel(
                f"[bold green]Real IPs:[/bold green] {', '.join(ips)}\n"
                f"[bold]Detection Methods:[/bold]\n" +
                "\n".join(f"  - {k}: {', '.join(v)}" for k, v in methods.items()),
                title="Real IP Detection",
                box=box.HEAVY,
                border_style="green",
            )
            console.print(panel)

    def print_port_scan(self, ports):
        if not ports:
            return
        table = Table(title="Open Ports", box=box.ROUNDED, header_style="bold red")
        table.add_column("Port", style="yellow")
        table.add_column("Service", style="cyan")
        table.add_column("Banner", style="white")
        for p in ports[:30]:
            table.add_row(str(p["port"]), p.get("service", ""), p.get("banner", "")[:60])
        console.print(table)

    def print_mail_security(self, mail_data):
        if not mail_data:
            return
        spf = mail_data.get("spf", {})
        dmarc = mail_data.get("dmarc", {})
        dkim = mail_data.get("dkim", {})
        panel = Panel(
            f"[bold]SPF:[/bold] {'[green]Present[/green]' if spf.get('present') else '[red]Missing[/red]'}\n"
            f"[bold]DKIM:[/bold] {'[green]Present[/green]' if dkim.get('present') else '[red]Missing[/red]'}\n"
            f"[bold]DMARC:[/bold] {'[green]Present[/green]' if dmarc.get('present') else '[red]Missing[/red]'} (Policy: {dmarc.get('policy', 'N/A')})",
            title="Mail Security (SPF/DKIM/DMARC)",
            box=box.HEAVY,
            border_style="cyan",
        )
        console.print(panel)

    def print_security_headers(self, headers_data):
        if not headers_data:
            return
        findings = headers_data.get("findings", [])
        if not findings:
            return
        table = Table(title="Security Headers Analysis", box=box.ROUNDED, header_style="bold white")
        table.add_column("Header", style="cyan")
        table.add_column("Present", style="yellow")
        table.add_column("Rating", style="red")
        for f in findings:
            present = "[green]Yes[/green]" if f.get("present") else "[red]No[/red]"
            rating_color = {"critical": "red", "high": "yellow", "medium": "blue"}.get(f.get("rating", ""), "white")
            table.add_row(f["header"], present, f"[{rating_color}]{f['rating']}[/{rating_color}]")
        console.print(table)

    def print_emails(self, emails):
        if not emails:
            return
        console.print("[bold red]Email Addresses Found[/bold red]")
        for e in emails:
            console.print(f"  [green]*[/green] [cyan]{e}[/cyan]")
        console.print()

    def print_takeover(self, takeover_results):
        if not takeover_results:
            return
        vulnerable = [t for t in takeover_results if t.get("vulnerable")]
        if not vulnerable:
            console.print("[green]No subdomain takeover vulnerabilities found.[/green]")
            return
        table = Table(title="Subdomain Takeover Vulnerabilities!", box=box.HEAVY, header_style="bold red")
        table.add_column("Hostname", style="yellow")
        table.add_column("Service", style="cyan")
        for v in vulnerable:
            table.add_row(v["hostname"], v.get("service", ""))
        console.print(table)

    def print_cve(self, cve_results):
        if not cve_results:
            return
        table = Table(title="CVE Findings", box=box.ROUNDED, header_style="bold red")
        table.add_column("Technology", style="yellow")
        table.add_column("Version", style="cyan")
        table.add_column("CVEs", style="white")
        for c in cve_results:
            table.add_row(c.get("technology", ""), c.get("version", ""), ", ".join(c.get("cves", [])))
        console.print(table)

    def print_cloud_results(self, cloud_data):
        if not cloud_data:
            return
        gcp = cloud_data.get("gcp", [])
        azure = cloud_data.get("azure", [])
        if gcp or azure:
            console.print(f"[bold cyan]Cloud Storage Scan[/bold cyan]")
            for bucket in gcp:
                console.print(f"  GCP: {bucket.get('url', '')} (accessible: {bucket.get('accessible', False)})")
            for container in azure:
                console.print(f"  Azure: {container.get('url', '')} (accessible: {container.get('accessible', False)})")

    def print_cms_results(self, cms_data):
        if not cms_data:
            return
        console.print("[bold yellow]CMS Detection[/bold yellow]")
        for cms, info in cms_data.items():
            if info.get("detected"):
                console.print(f"  [green]*[/green] {cms} v{info.get('version', '?')} (confidence: {info.get('confidence', 0)}%)")

    def print_waf_results(self, waf_data):
        if waf_data and waf_data.get("waf_detected"):
            console.print(f"[bold yellow]WAF Detected: {waf_data.get('waf_name')} ({waf_data.get('confidence')}%)[/bold yellow]")
        else:
            console.print("[green]No WAF detected.[/green]")

    def print_cicd_results(self, cicd_data):
        if not cicd_data:
            return
        console.print(f"[bold cyan]CI/CD Files Found: {len(cicd_data)}[/bold cyan]")
        for item in cicd_data[:10]:
            console.print(f"  {item.get('url', '')} (status: {item.get('status', '')})")

    def print_swagger_results(self, swagger_data, graphql_data):
        if swagger_data:
            console.print(f"[bold cyan]API Documentation Found[/bold cyan]")
            for item in swagger_data:
                console.print(f"  {item.get('url', '')} (type: {item.get('type', '')})")
        if graphql_data:
            console.print(f"[bold yellow]GraphQL Endpoints Found[/bold yellow]")
            for item in graphql_data:
                console.print(f"  {item.get('url', '')} (introspection: {item.get('introspection', False)})")

    def print_docker_results(self, docker_data):
        if not docker_data:
            return
        if docker_data.get("exposed"):
            console.print("[bold red]Docker Registry Exposed![/bold red]")
            for reg in docker_data.get("registries", []):
                console.print(f"  {reg.get('url', '')}")
        else:
            console.print("[green]No exposed Docker registries.[/green]")

    def _collapsible(self, title, content, badge="", open_default=False):
        open_attr = " open" if open_default else ""
        badge_html = f'<span class="badge">{badge}</span>' if badge else ""
        return f'''<details class="collapsible"{open_attr}>
        <summary>{title}{badge_html}</summary>
        <div class="section-content">{content}</div>
    </details>'''

    def _build_dns_records_section(self):
        d = self.data.get("dns", {})
        if not d:
            return ""
        rows = ""
        for hostname, records in d.items():
            if not records:
                continue
            record_lines = ""
            has_soa = False
            for rtype in ["A", "AAAA", "MX", "NS", "TXT", "CNAME", "SOA"]:
                if rtype not in records:
                    continue
                vals = records[rtype]
                if rtype == "SOA":
                    has_soa = True
                    soa = vals
                    record_lines += f'<div class="dr"><span class="rt">SOA</span><span class="rv">{soa.get("mname", "")} — serial {soa.get("serial", "")}</span></div>'
                elif isinstance(vals, list):
                    for v in vals[:5]:
                        record_lines += f'<div class="dr"><span class="rt">{rtype}</span><span class="rv">{v}</span></div>'
            if not has_soa and "SOA" not in records:
                pass
            if record_lines:
                rows += f'<tr><td class="dh">{hostname}</td><td class="dr-cell">{record_lines}</td></tr>'
        if not rows:
            return ""
        content = f'<table class="dns-records-table"><tr><th>Hostname</th><th>Records</th></tr>{rows}</table>'
        return self._collapsible("DNS Records", content, str(len(d)))

    def _build_ssl_section(self):
        d = self.data.get("ssl", {})
        if not d:
            return ""
        cards = ""
        for hostname, cert in d.items():
            issuer = ""
            for k in ["common_name", "name", "organization", "organizational_unit"]:
                v = cert.get("issuer", {}).get(k, "")
                if v:
                    issuer = v
                    break
            subject = ""
            for k in ["common_name", "name", "organization"]:
                v = cert.get("subject", {}).get(k, "")
                if v:
                    subject = v
                    break
            if not subject:
                subject = hostname
            sans = ", ".join(cert.get("alt_names", [])[:10]) or "N/A"
            fp = cert.get("fingerprint", "N/A")
            if len(fp) > 32:
                fp = fp[:32] + "…"
            vf = str(cert.get("valid_from", ""))[:10]
            vt = str(cert.get("valid_to", ""))[:10]
            days = cert.get("days_remaining", "?")
            expired = cert.get("is_expired", False)
            sc = "expired" if expired else "valid"
            cards += f'''<div class="ssl-card {sc}">
                <div class="ssl-host">{hostname}</div>
                <div class="ssl-subj">Subject: {subject}</div>
                <div class="ssl-d"><span>Issuer:</span> {issuer}</div>
                <div class="ssl-d"><span>Valid:</span> {vf} → {vt} <span class="days-{sc}">({days} days)</span></div>
                <div class="ssl-d"><span>FP:</span> <code>{fp}</code></div>
                <div class="ssl-d"><span>SANs:</span> <span class="sans">{sans}</span></div>
            </div>'''
        if not cards:
            return ""
        return self._collapsible("SSL Certificates", cards, str(len(d)))

    def _build_asn_section(self):
        d = self.data.get("asn_geo", {})
        if not d or not d.get("asn"):
            return ""
        items = [
            ("ASN", d.get("asn", "N/A")),
            ("Country", d.get("country", "N/A")),
            ("City", d.get("city", "N/A")),
            ("ISP", d.get("isp", "N/A")),
            ("Organization", d.get("org", "N/A")),
            ("Latitude", str(d.get("latitude", "N/A"))),
            ("Longitude", str(d.get("longitude", "N/A"))),
        ]
        rows = "".join(f"<tr><td style='color:#5a7a9a;font-weight:600;'>{k}</td><td>{v}</td></tr>" for k, v in items)
        content = f'<table class="asn-table">{rows}</table>'
        return f'<div class="section"><h2>ASN / Geo</h2>{content}</div>'

    def _build_emails_section(self):
        d = self.data.get("emails", [])
        if not d:
            return ""
        items = "".join(f'<div class="finding-item"><span style="color:#00ff88;">✉</span> <span class="endpoint-link">{e}</span></div>' for e in d[:50])
        if len(d) > 50:
            items += f'<div style="color:#5a7a9a;margin-top:8px;">… and {len(d)-50} more</div>'
        return self._collapsible("Emails Found", items, str(len(d)))

    def _build_js_section(self):
        d = self.data.get("js_endpoints", {})
        if not d:
            return ""
        items = ""
        count = sum(len(v) for v in d.values())
        for hostname, endpoints in list(d.items())[:10]:
            items += f'<div style="margin-bottom:8px;"><span style="color:#00d4ff;font-weight:600;">{hostname}</span></div>'
            for ep in endpoints[:15]:
                items += f'<div class="dr"><span class="rt">JS</span><span class="rv endpoint-link">{ep}</span></div>'
            if len(endpoints) > 15:
                items += f'<div style="color:#5a7a9a;margin-left:20px;">… +{len(endpoints)-15} more</div>'
        if not items:
            return ""
        return self._collapsible("JS Endpoints", items, str(count))

    def _build_wayback_section(self):
        d = self.data.get("wayback_urls", [])
        if not d:
            return ""
        items = "".join(f'<div class="dr"><span class="rt">WB</span><span class="rv endpoint-link">{u}</span></div>' for u in d[:30])
        if len(d) > 30:
            items += f'<div style="color:#5a7a9a;margin-top:8px;">… and {len(d)-30} more</div>'
        return self._collapsible("Wayback Machine URLs", items, str(len(d)))

    def _build_commoncrawl_section(self):
        d = self.data.get("commoncrawl_urls", [])
        if not d:
            return ""
        items = "".join(f'<div class="dr"><span class="rt">CC</span><span class="rv endpoint-link">{u}</span></div>' for u in d[:30])
        if len(d) > 30:
            items += f'<div style="color:#5a7a9a;margin-top:8px;">… and {len(d)-30} more</div>'
        return self._collapsible("CommonCrawl URLs", items, str(len(d)))

    def _build_dorks_section(self):
        d = self.data.get("google_dorks", [])
        if not d:
            return ""
        items = "".join(f'<div class="finding-item"><span style="color:#ffd93d;">⌕</span> <code style="color:#c8d6e5;font-size:12px;">{dk}</code></div>' for dk in d[:40])
        if len(d) > 40:
            items += f'<div style="color:#5a7a9a;margin-top:8px;">… and {len(d)-40} more</div>'
        return self._collapsible("Google Dorks", items, str(len(d)))

    def _build_github_section(self):
        d = self.data.get("github", {})
        if not d:
            return ""
        repos = d.get("repos", [])
        findings = d.get("findings", [])
        items = ""
        for r in repos[:15]:
            items += f'<div class="finding-item"><span style="color:#6e5494;">⬡</span> <span class="endpoint-link">{r.get("url", r.get("name", ""))}</span></div>'
        for f in findings[:15]:
            items += f'<div class="finding-item"><span style="color:#ff4444;">⚠</span> <span>{f.get("message", str(f))}</span></div>'
        if not items:
            return ""
        total = len(repos) + len(findings)
        return self._collapsible("GitHub Recon", items, str(total))

    def _build_sec_headers_section(self):
        d = self.data.get("security_headers", {})
        if not d:
            return ""
        findings = d.get("findings", [])
        if not findings:
            return ""
        rows = ""
        for f in findings:
            present = f.get("present", False)
            cls = "present" if present else "missing"
            rating = f.get("rating", "info")
            rcolor = {"critical": "#ff4444", "high": "#ffd93d", "medium": "#00d4ff", "info": "#5a7a9a"}.get(rating, "#5a7a9a")
            rows += f'''<tr>
                <td>{f.get("header", "")}</td>
                <td><span class="finding-status {cls}"></span></td>
                <td style="color:{rcolor};">{rating}</td>
            </tr>'''
        if not rows:
            return ""
        content = f'<table class="sec-headers-table"><tr><th>Header</th><th>Status</th><th>Risk</th></tr>{rows}</table>'
        return self._collapsible("Security Headers", content, str(len(findings)))

    def _build_cors_section(self):
        d = self.data.get("cors", {})
        if not d:
            return ""
        items = ""
        if d.get("wildcard"):
            items += '<div class="finding-item"><span class="finding-status present"></span> <span style="color:#ff4444;">Wildcard CORS origin allowed</span></div>'
        if d.get("origin_reflection"):
            items += '<div class="finding-item"><span class="finding-status present"></span> <span style="color:#ffd93d;">Origin reflection detected</span></div>'
        if d.get("credentials"):
            items += '<div class="finding-item"><span class="finding-status info"></span> <span>Credentials allowed: {}</span></div>'.format(d["credentials"])
        if not items:
            items = '<div class="finding-item"><span class="finding-status info"></span> No CORS misconfigurations found</div>'
        return self._collapsible("CORS Analysis", items)

    def _build_cookies_section(self):
        d = self.data.get("cookies", {})
        if not d:
            return ""
        items = ""
        for hostname, cookies in d.items():
            for c in (cookies if isinstance(cookies, list) else [cookies]):
                if isinstance(c, dict):
                    name = c.get("name", str(c))
                    secure = c.get("secure", False)
                    httponly = c.get("httponly", False)
                    samesite = c.get("samesite", "N/A")
                    items += f'<div class="finding-item">'
                    items += f'<span class="finding-status {"present" if secure else "missing"}"></span>'
                    items += f'<span style="font-weight:600;">{name}</span> '
                    items += f'<span style="color:#5a7a9a;">Secure:{"✅" if secure else "❌"} HttpOnly:{"✅" if httponly else "❌"} SameSite:{samesite}</span>'
                    items += f'</div>'
        if not items:
            return ""
        return self._collapsible("Cookie Analysis", items)

    def _build_favicon_section(self):
        d = self.data.get("favicon", {})
        if not d or not d.get("mmh3"):
            return ""
        content = f'''<div style="display:flex;align-items:center;gap:16px;">
            <div><span style="color:#5a7a9a;">mmh3:</span> <code style="color:#00d4ff;">{d.get("mmh3", "")}</code></div>
            <div><span style="color:#5a7a9a;">Size:</span> {d.get("size", "?")} bytes</div>
            <div><span style="color:#5a7a9a;">URL:</span> <span class="endpoint-link">{d.get("url", "N/A")}</span></div>
        </div>'''
        return f'<div class="section"><h2>Favicon Hash</h2>{content}</div>'

    def _build_screenshots_section(self):
        d = self.data.get("screenshots", [])
        if not d:
            return ""
        items = ""
        for s in d[:10]:
            path = s.get("path", s.get("filepath", ""))
            url = s.get("url", "")
            if path and os.path.exists(path):
                import base64
                try:
                    with open(path, "rb") as f:
                        b64 = base64.b64encode(f.read()).decode()
                    ext = os.path.splitext(path)[1].lower()
                    if ext in [".png", ".jpg", ".jpeg", ".gif", ".webp"]:
                        mime = "image/png" if ext == ".png" else "image/jpeg"
                        items += f'<div style="margin-bottom:12px;"><div style="color:#00d4ff;font-size:12px;margin-bottom:4px;">{url}</div><img src="data:{mime};base64,{b64}" style="max-width:100%;border-radius:6px;border:1px solid #1a2a4a;"></div>'
                    else:
                        items += f'<div class="finding-item"><span class="endpoint-link">{url}</span> → {path}</div>'
                except Exception:
                    items += f'<div class="finding-item"><span class="endpoint-link">{url}</span> → {path}</div>'
            else:
                items += f'<div class="finding-item"><span class="endpoint-link">{url}</span> → {path}</div>'
        if not items:
            return ""
        return self._collapsible("Screenshots", items, str(len(d)))

    def _build_spider_section(self):
        d = self.data.get("spider_links", {})
        if not d:
            return ""
        items = ""
        count = sum(len(v) for v in d.values())
        for hostname, links in list(d.items())[:5]:
            items += f'<div style="margin-bottom:6px;"><span style="color:#00d4ff;font-weight:600;">{hostname}</span> <span style="color:#5a7a9a;">({len(links)} links)</span></div>'
            for link in links[:20]:
                items += f'<div class="dr"><span class="rt">→</span><span class="rv endpoint-link">{link if isinstance(link, str) else link.get("url", str(link))}</span></div>'
            if len(links) > 20:
                items += f'<div style="color:#5a7a9a;margin-left:20px;">… +{len(links)-20} more</div>'
        if not items:
            return ""
        return self._collapsible("Spider Links", items, str(count))

    def _build_s3_section(self):
        d = self.data.get("s3", [])
        if not d:
            return ""
        items = ""
        for b in d[:20]:
            accessible = b.get("accessible", False)
            url = b.get("url", b.get("name", ""))
            items += f'<div class="finding-item"><div class="finding-status {"present" if accessible else "info"}"></div><span class="endpoint-link">{url}</span></div>'
        if not items:
            return ""
        return self._collapsible("S3 Buckets", items, str(len(d)))

    def _build_dirs_section(self):
        d = self.data.get("directory_enum", [])
        if not d:
            return ""
        rows = ""
        for p in d[:50]:
            status = p.get("status", "")
            size = p.get("size", "")
            url = p.get("url", "")
            sc = "alive" if status and status < 400 else "dead"
            rows += f'<tr><td class="endpoint-link">{url}</td><td class="{sc}">{status}</td><td>{size}</td></tr>'
        if not rows:
            return ""
        content = f'<table class="dirs-table"><tr><th>URL</th><th>Status</th><th>Size</th></tr>{rows}</table>'
        if len(d) > 50:
            content += f'<div style="color:#5a7a9a;margin-top:8px;">… and {len(d)-50} more</div>'
        return self._collapsible("Directory Enumeration", content, str(len(d)))

    def _build_findings_section(self):
        if not self._findings:
            return ""
        rows = ""
        sev_icon = {"critical": "!!", "high": "!", "medium": "~", "info": "i"}
        sev_color = {"critical": "#ff4444", "high": "#ffd93d", "medium": "#00d4ff", "info": "#5a7a9a"}
        sev_bg = {"critical": "rgba(255,68,68,0.08)", "high": "rgba(255,217,61,0.08)", "medium": "rgba(0,212,255,0.08)", "info": "rgba(90,122,154,0.08)"}
        for f in self._findings:
            sev = f.get("severity", "info")
            rows += f'''<tr style="background:{sev_bg.get(sev, 'transparent')};">
                <td><span style="color:{sev_color.get(sev, '#5a7a9a')};font-weight:700;">{sev_icon.get(sev, "?")}</span></td>
                <td style="color:{sev_color.get(sev, '#5a7a9a')};font-weight:600;">{sev.upper()}</td>
                <td style="color:#5a7a9a;font-size:11px;">{f.get("category", "")}</td>
                <td>{f.get("message", "")}</td>
            </tr>'''
        if not rows:
            return ""
        content = f'<table class="findings-table"><tr><th></th><th>Severity</th><th>Category</th><th>Message</th></tr>{rows}</table>'
        return self._collapsible("Vulnerability Findings", content, str(len(self._findings)))

    def _build_html_sections(self):
        d = self.data

        subdomains_table_rows = ""
        for s in d.get("subdomains", [])[:200]:
            hostname = s if isinstance(s, str) else s.get("hostname", "")
            ip = ""
            status = ""
            server = ""
            title = ""
            status_class = "unknown"

            if isinstance(s, dict):
                ip = s.get("ip", "")
                status = s.get("status", "")
                server = s.get("server", "")
                title = s.get("title", "")

            dns_entry = d.get("dns", {}).get(hostname, {})
            if dns_entry and not ip:
                ips = dns_entry.get("A", [])
                if ips:
                    ip = ips[0]
                elif dns_entry.get("AAAA"):
                    ip = dns_entry["AAAA"][0]

            http_entry = d.get("http", {}).get(hostname, {})
            if http_entry:
                sc = http_entry.get("status_code", "")
                alive = http_entry.get("alive", False)
                if sc:
                    status = str(sc)
                    status_class = "alive" if alive else "dead"
                server = http_entry.get("server", server)
                title = http_entry.get("title", title)

            if not ip:
                ip_display = '<span class="no-dns">No DNS</span>'
            else:
                ip_display = ip
            if not status:
                status_display = '<span class="no-http">—</span>'
            else:
                status_display = status

            subdomains_table_rows += f'''<tr>
                <td>{hostname}</td>
                <td>{ip_display}</td>
                <td class="{status_class}">{status_display}</td>
                <td>{server[:30]}</td>
                <td>{title[:60]}</td>
            </tr>'''

        whois = d.get("whois", {})
        whois_status = "ACTIVE" if whois.get("is_registered") and not whois.get("is_expired") else "EXPIRED" if whois.get("is_expired") else "UNKNOWN"
        whois_html = ""
        if whois and (whois.get("registrar") or whois.get("is_registered")):
            status_class = "danger" if whois_status == "EXPIRED" else "success"
            whois_html = f"""<div class="section">
                <h2>WHOIS Information</h2>
                <table class="whois-table"><tr><td style="color:#5a7a9a;font-weight:600;">Registrar</td><td>{whois.get("registrar", "N/A")}</td></tr>
                <tr><td style="color:#5a7a9a;font-weight:600;">Created</td><td>{whois.get("created", "N/A")}</td></tr>
                <tr><td style="color:#5a7a9a;font-weight:600;">Expires</td><td>{whois.get("expires", "N/A")}</td></tr>
                <tr><td style="color:#5a7a9a;font-weight:600;">Status</td><td class="{status_class}">{whois_status}</td></tr>
                <tr><td style="color:#5a7a9a;font-weight:600;">Days Until Expiry</td><td>{whois.get("days_until_expiry", "N/A")}</td></tr></table>
            </div>"""

        real_ip = d.get("real_ip", {})
        real_ip_html = ""
        if real_ip and real_ip.get("ips"):
            ips = ", ".join(real_ip["ips"])
            methods_str = ", ".join(real_ip.get("methods", {}).keys())
            real_ip_html = f"""<div class="section">
                <h2>Real IP Detection</h2>
                <div class="real-ip">{ips}</div>
                <p style="margin-top:10px;color:#5a7a9a;font-size:13px;">Methods: {methods_str}</p>
            </div>"""

        ports = d.get("port_scan", {}).get("results", [])
        ports_html = ""
        if ports:
            rows = "".join(f"<tr><td>{p['port']}</td><td>{p.get('service', '')}</td><td>{p.get('banner', '')[:50]}</td></tr>" for p in ports[:30])
            ports_html = f"""<div class="section {'danger' if ports else ''}"><h2>Open Ports <span class="badge {'danger' if ports else 'success'}">{len(ports)}</span></h2><table><tr><th>Port</th><th>Service</th><th>Banner</th></tr>{rows}</table></div>"""

        mail = d.get("mail_security", {})
        mail_html = ""
        if mail:
            spf_status = "present" if mail.get("spf", {}).get("present") else "missing"
            dkim_status = "present" if mail.get("dkim", {}).get("present") else "missing"
            dmarc_status = "present" if mail.get("dmarc", {}).get("present") else "missing"
            mail_html = f"""<div class="section"><h2>Mail Security</h2>
                <div class="finding-item"><div class="finding-status {spf_status}"></div><span style="font-weight:600;">SPF:</span> <span style="color:{'#00ff88' if spf_status == 'present' else '#ff4444'};">{'Present' if spf_status == 'present' else 'Missing'}</span></div>
                <div class="finding-item"><div class="finding-status {dkim_status}"></div><span style="font-weight:600;">DKIM:</span> <span style="color:{'#00ff88' if dkim_status == 'present' else '#ff4444'};">{'Present' if dkim_status == 'present' else 'Missing'}</span></div>
                <div class="finding-item"><div class="finding-status {dmarc_status}"></div><span style="font-weight:600;">DMARC:</span> <span style="color:{'#00ff88' if dmarc_status == 'present' else '#ff4444'};">{'Present' if dmarc_status == 'present' else 'Missing'}</span> <span style="color:#5a7a9a;">(Policy: {mail.get('dmarc', {}).get('policy', 'N/A')})</span></div>
            </div>"""

        tech = d.get("technology", {})
        tech_html = ""
        if tech:
            flat_tech = []
            for sub, tlist in tech.items():
                if isinstance(tlist, dict):
                    flat_tech.extend(list(tlist.keys()))
            unique_tech = list(set(flat_tech))[:30]
            if unique_tech:
                tags = "".join(f'<span class="tech-tag">{t}</span>' for t in sorted(unique_tech))
                tech_html = f"""<div class="section"><h2>Technology Stack <span class="badge">{len(unique_tech)}</span></h2><div class="tech-list">{tags}</div></div>"""

        cve = d.get("cve", [])
        cve_html = ""
        if cve:
            def _cve_row(c):
                cves = c.get("cves", [])
                tags = " ".join(f'<span class="cve-tag">{cv}</span>' for cv in cves)
                return f"<tr><td>{c.get('technology', '')}</td><td>{c.get('version', '')}</td><td>{tags}</td></tr>"
            rows = "".join(_cve_row(c) for c in cve)
            cve_html = f"""<div class="section vuln-section"><h2>CVE Findings <span class="badge danger">{len(cve)}</span></h2><table><tr><th>Technology</th><th>Version</th><th>CVEs</th></tr>{rows}</table></div>"""

        takeover = d.get("takeover", [])
        takeover_html = ""
        if takeover:
            vuln = [t for t in takeover if t.get("vulnerable")]
            if vuln:
                rows = "".join(f"<tr><td>{v['hostname']}</td><td>{v.get('service', '')}</td></tr>" for v in vuln)
                takeover_html = f"""<div class="section vuln-section"><h2>Subdomain Takeover Vulnerabilities <span class="badge danger">{len(vuln)}</span></h2><table><tr><th>Hostname</th><th>Service</th></tr>{rows}</table></div>"""

        cloud = d.get("cloud_storage", {})
        cloud_html = ""
        if cloud:
            gcp = cloud.get("gcp", [])
            azure = cloud.get("azure", [])
            items = []
            for b in gcp:
                items.append(f'<div class="finding-item"><div class="finding-status {"present" if b.get("accessible") else "info"}"></div><span class="endpoint-link">{b.get("url", "")}</span> GCP bucket</div>')
            for c in azure:
                items.append(f'<div class="finding-item"><div class="finding-status {"present" if c.get("accessible") else "info"}"></div><span class="endpoint-link">{c.get("url", "")}</span> Azure container</div>')
            if items:
                cloud_html = f"""<div class="section"><h2>Cloud Storage</h2>{"".join(items)}</div>"""

        cms = d.get("cms", {})
        cms_html = ""
        if cms:
            items = []
            for name, info in cms.items():
                if info.get("detected"):
                    items.append(f'<span class="tech-tag">{name} v{info.get("version", "?")} ({info.get("confidence", 0)}%)</span>')
            if items:
                cms_html = f"""<div class="section"><h2>CMS Detection <span class="badge">{len(items)}</span></h2><div class="tech-list">{"".join(items)}</div></div>"""

        waf = d.get("waf", {})
        waf_html = ""
        if waf:
            if waf.get("waf_detected"):
                waf_html = f"""<div class="section vuln-section"><h2>WAF Detection</h2><div class="finding-item"><div class="finding-status present"></div><span style="font-weight:600;">{waf.get("waf_name", "Unknown")}</span> <span style="color:#5a7a9a;">(confidence: {waf.get("confidence", 0)}%)</span></div></div>"""
            else:
                waf_html = f"""<div class="section"><h2>WAF Detection</h2><div class="finding-item"><div class="finding-status info"></div>No WAF detected</div></div>"""

        cicd = d.get("cicd", [])
        cicd_html = ""
        if cicd:
            items = "".join(f'<div class="finding-item"><div class="finding-status present"></div><span class="endpoint-link">{c.get("url", "")}</span> <span style="color:#5a7a9a;">status {c.get("status", "")}</span></div>' for c in cicd[:15])
            cicd_html = f"""<div class="section"><h2>CI/CD Files <span class="badge">{len(cicd[:15])}</span></h2>{items}</div>"""

        swag = d.get("swagger", [])
        gql = d.get("graphql", [])
        api_html = ""
        if swag:
            items = "".join(f'<div class="finding-item"><div class="finding-status info"></div><span class="endpoint-link">{s.get("url", "")}</span> <span style="color:#5a7a9a;">({s.get("type", "")})</span></div>' for s in swag)
            api_html += f"""<div class="section"><h2>API Documentation <span class="badge">{len(swag)}</span></h2>{items}</div>"""
        if gql:
            items = "".join(f'<div class="finding-item"><div class="finding-status {"danger" if g.get("introspection") else "info"}"></div><span class="endpoint-link">{g.get("url", "")}</span> <span style="color:#5a7a9a;">introspection: {g.get("introspection", False)}</span></div>' for g in gql)
            api_html += f"""<div class="section"><h2>GraphQL Endpoints <span class="badge">{len(gql)}</span></h2>{items}</div>"""

        docker = d.get("docker_registry", {})
        docker_html = ""
        if docker and docker.get("exposed"):
            items = "".join(f'<div class="finding-item"><div class="finding-status danger"></div><span class="endpoint-link">{r.get("url", "")}</span></div>' for r in docker.get("registries", []))
            docker_html = f"""<div class="section vuln-section"><h2>Docker Registry Exposed <span class="badge danger">{len(docker.get("registries", []))}</span></h2>{items}</div>"""

        dns_records_html = self._build_dns_records_section()
        ssl_html = self._build_ssl_section()
        asn_html = self._build_asn_section()
        emails_html = self._build_emails_section()
        js_html = self._build_js_section()
        wayback_html = self._build_wayback_section()
        commoncrawl_html = self._build_commoncrawl_section()
        dorks_html = self._build_dorks_section()
        github_html = self._build_github_section()
        sec_headers_html = self._build_sec_headers_section()
        cors_html = self._build_cors_section()
        cookies_html = self._build_cookies_section()
        favicon_html = self._build_favicon_section()
        screenshots_html = self._build_screenshots_section()
        spider_html = self._build_spider_section()
        s3_html = self._build_s3_section()
        dirs_html = self._build_dirs_section()
        findings_html = self._build_findings_section()

        return {
            "subdomains_table_rows": subdomains_table_rows,
            "whois_html": whois_html,
            "whois_status": whois_status,
            "real_ip_html": real_ip_html,
            "ports_html": ports_html,
            "mail_html": mail_html,
            "tech_html": tech_html,
            "cve_html": cve_html,
            "takeover_html": takeover_html,
            "cloud_html": cloud_html,
            "cms_html": cms_html,
            "waf_html": waf_html,
            "cicd_html": cicd_html,
            "api_html": api_html,
            "docker_html": docker_html,
            "dns_records_html": dns_records_html,
            "ssl_html": ssl_html,
            "asn_html": asn_html,
            "emails_html": emails_html,
            "js_html": js_html,
            "wayback_html": wayback_html,
            "commoncrawl_html": commoncrawl_html,
            "dorks_html": dorks_html,
            "github_html": github_html,
            "sec_headers_html": sec_headers_html,
            "cors_html": cors_html,
            "cookies_html": cookies_html,
            "favicon_html": favicon_html,
            "screenshots_html": screenshots_html,
            "spider_html": spider_html,
            "s3_html": s3_html,
            "dirs_html": dirs_html,
            "findings_html": findings_html,
        }

    def export_json(self, filepath):
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=2, default=str)
        console.print(f"[green]JSON exported to: {filepath}[/green]")

    def export_yaml(self, filepath):
        try:
            import yaml
            with open(filepath, "w", encoding="utf-8") as f:
                yaml.dump(self.data, f, default_flow_style=False, allow_unicode=True, sort_keys=False)
            console.print(f"[green]YAML exported to: {filepath}[/green]")
        except ImportError:
            self.export_json(filepath.replace(".yaml", ".json").replace(".yml", ".json"))
            console.print("[yellow]PyYAML not installed. JSON exported instead.[/yellow]")

    def export_csv(self, filepath):
        subdomains = self.data.get("subdomains", [])
        if not subdomains:
            console.print("[red]No data to export to CSV.[/red]")
            return
        fieldnames = ["hostname", "ip", "status", "server", "title", "cdn"]
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for s in subdomains:
                row = {
                    "hostname": s if isinstance(s, str) else s.get("hostname", ""),
                    "ip": "",
                    "status": "",
                    "server": "",
                    "title": "",
                    "cdn": "",
                }
                if isinstance(s, dict):
                    row.update({k: s.get(k, "") for k in fieldnames if k in s})
                writer.writerow(row)
        console.print(f"[green]CSV exported to: {filepath}[/green]")

    def export_xlsx(self, filepath):
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Subdomains"
            ws.append(["Hostname", "IP", "Status", "Server", "Title", "CDN"])
            for s in self.data.get("subdomains", []):
                if isinstance(s, str):
                    ws.append([s, "", "", "", "", ""])
                elif isinstance(s, dict):
                    ws.append([
                        s.get("hostname", ""),
                        s.get("ip", ""),
                        s.get("status", ""),
                        s.get("server", ""),
                        s.get("title", ""),
                        s.get("cdn", ""),
                    ])
            ws2 = wb.create_sheet("WHOIS")
            whois = self.data.get("whois", {})
            for k, v in whois.items():
                if isinstance(v, list):
                    v = ", ".join(str(x) for x in v)
                ws2.append([str(k), str(v)])
            ws3 = wb.create_sheet("Ports")
            ws3.append(["Port", "Service", "Banner"])
            for p in self.data.get("port_scan", {}).get("results", []):
                ws3.append([p.get("port", ""), p.get("service", ""), p.get("banner", "")])
            wb.save(filepath)
            console.print(f"[green]XLSX exported to: {filepath}[/green]")
        except ImportError:
            console.print("[yellow]openpyxl not installed. Skipping XLSX export.[/yellow]")

    def export_html(self, filepath):
        sections = self._build_html_sections()
        d = self.data
        subs_count = len(d.get("subdomains", []))
        emails_count = len(d.get("emails", []))
        ports_count = len(d.get("port_scan", {}).get("results", []))
        alive_count = sum(1 for h in d.get("http", {}).values() if h.get("alive"))
        takeover_count = len(d.get("takeover", []))
        dir_count = len(d.get("directory_enum", []))
        tech_count = len(d.get("technology", {}))
        cve_count = len(d.get("cve", []))
        whois_status = sections['whois_status']

        critical_count = len([f for f in getattr(self, '_findings', []) if f.get('severity') == 'critical'])
        high_count = len([f for f in getattr(self, '_findings', []) if f.get('severity') == 'high'])
        medium_count = len([f for f in getattr(self, '_findings', []) if f.get('severity') == 'medium'])

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>XREFS0 Recon Report — {self.domain}</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&family=JetBrains+Mono:wght@400;600&display=swap');
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ background: #080c18; color: #c8d6e5; font-family: 'Inter', system-ui, -apple-system, sans-serif; line-height: 1.6; }}
::-webkit-scrollbar {{ width: 8px; }}
::-webkit-scrollbar-track {{ background: #0a0e1a; }}
::-webkit-scrollbar-thumb {{ background: #1a2a4a; border-radius: 4px; }}
::-webkit-scrollbar-thumb:hover {{ background: #2a3a5a; }}

.header {{ background: linear-gradient(135deg, #0a0e1a 0%, #14142a 50%, #1a0a1a 100%); border-bottom: 1px solid #1a2a4a; padding: 40px 48px; position: relative; overflow: hidden; }}
.header::before {{ content: ''; position: absolute; inset: 0; background: radial-gradient(ellipse at 20% 50%, rgba(0,212,255,0.06) 0%, transparent 60%), radial-gradient(ellipse at 80% 50%, rgba(255,68,68,0.04) 0%, transparent 60%); }}
.header-content {{ position: relative; z-index: 1; display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 16px; }}
.header h1 {{ font-size: 28px; font-weight: 800; color: #00d4ff; letter-spacing: 2px; }}
.header h1 .dot {{ color: #ff4444; }}
.header .meta {{ color: #5a7a9a; font-size: 13px; }}
.header .meta strong {{ color: #00d4ff; }}
.header .badge-group {{ display: flex; gap: 8px; flex-wrap: wrap; }}
.header .badge {{ padding: 4px 14px; border-radius: 6px; font-size: 11px; font-weight: 600; letter-spacing: 1px; text-transform: uppercase; }}
.header .badge.critical {{ background: rgba(255,68,68,0.15); color: #ff4444; border: 1px solid rgba(255,68,68,0.3); }}
.header .badge.high {{ background: rgba(255,217,61,0.12); color: #ffd93d; border: 1px solid rgba(255,217,61,0.3); }}
.header .badge.info {{ background: rgba(0,212,255,0.1); color: #00d4ff; border: 1px solid rgba(0,212,255,0.2); }}

.container {{ max-width: 1440px; margin: 0 auto; padding: 32px 48px; }}

.stats-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(140px, 1fr)); gap: 12px; margin-bottom: 32px; }}
.stat-card {{ background: #0d1520; border: 1px solid #1a2a4a; border-radius: 10px; padding: 18px 14px; text-align: center; transition: border-color 0.2s, transform 0.2s; }}
.stat-card:hover {{ border-color: #2a3a5a; transform: translateY(-1px); }}
.stat-card .value {{ font-size: 30px; font-weight: 800; color: #00d4ff; line-height: 1.2; }}
.stat-card .label {{ font-size: 10px; color: #5a7a9a; text-transform: uppercase; letter-spacing: 1.5px; margin-top: 4px; }}
.stat-card.danger .value {{ color: #ff4444; }}
.stat-card.warning .value {{ color: #ffd93d; }}
.stat-card.success .value {{ color: #00ff88; }}

@media (max-width: 768px) {{ .container {{ padding: 16px 20px; }} .header {{ padding: 24px 20px; }} .header h1 {{ font-size: 22px; }} }}

.section {{ background: #0d1520; border: 1px solid #1a2a4a; border-radius: 10px; padding: 20px 24px; margin-bottom: 20px; }}
.section.danger {{ border-color: #ff4444; background: linear-gradient(135deg, #140a0a 0%, #0d1520 100%); }}
.section h2 {{ font-size: 15px; font-weight: 700; color: #00d4ff; margin-bottom: 14px; padding-bottom: 10px; border-bottom: 1px solid #1a2a4a; display: flex; align-items: center; gap: 10px; }}
.section.danger h2 {{ color: #ff4444; }}
.section h2 .badge {{ background: #00d4ff; color: #0a0e1a; font-size: 10px; font-weight: 700; padding: 1px 10px; border-radius: 12px; }}
.section h2 .badge.danger {{ background: #ff4444; color: #fff; }}
.section h2 .badge.success {{ background: #00ff88; color: #0a0e1a; }}

table {{ width: 100%; border-collapse: separate; border-spacing: 0; font-size: 13px; }}
th {{ padding: 10px 12px; text-align: left; font-size: 10px; font-weight: 700; color: #5a7a9a; text-transform: uppercase; letter-spacing: 1px; border-bottom: 1px solid #1a2a4a; }}
td {{ padding: 8px 12px; border-bottom: 1px solid rgba(26,42,74,0.4); }}
tr:hover td {{ background: rgba(0,212,255,0.02); }}
.alive {{ color: #00ff88; font-weight: 600; }}
.dead {{ color: #5a7a9a; }}

.finding-item {{ display: flex; align-items: flex-start; gap: 10px; padding: 8px 0; border-bottom: 1px solid rgba(26,42,74,0.3); font-size: 13px; }}
.finding-item:last-child {{ border: none; }}

.tech-list {{ display: flex; flex-wrap: wrap; gap: 6px; }}
.tech-tag {{ background: rgba(0,212,255,0.08); border: 1px solid rgba(0,212,255,0.15); color: #00d4ff; padding: 3px 12px; border-radius: 4px; font-size: 11px; font-weight: 500; }}
.cve-tag {{ background: rgba(255,68,68,0.08); border: 1px solid rgba(255,68,68,0.15); color: #ff6b6b; padding: 2px 10px; border-radius: 4px; font-size: 10px; }}
.finding-status {{ width: 10px; height: 10px; border-radius: 50%; flex-shrink: 0; }}
.finding-status.present {{ background: #00ff88; box-shadow: 0 0 6px rgba(0,255,136,0.4); }}
.finding-status.missing {{ background: #ff4444; }}
.finding-status.info {{ background: #00d4ff; }}
.no-dns {{ color: #5a7a9a; font-style: italic; }}
.no-http {{ color: #5a7a9a; font-style: italic; }}
.real-ip {{ font-size: 18px; font-weight: 700; color: #00ff88; }}

/* Collapsible sections */
details.collapsible {{ background: #0d1520; border: 1px solid #1a2a4a; border-radius: 10px; margin-bottom: 16px; overflow: hidden; transition: border-color 0.2s; }}
details.collapsible[open] {{ border-color: #2a3a5a; }}
details.collapsible summary {{ padding: 16px 24px; cursor: pointer; font-size: 15px; font-weight: 700; color: #00d4ff; display: flex; align-items: center; gap: 12px; user-select: none; position: relative; list-style: none; }}
details.collapsible summary::-webkit-details-marker {{ display: none; }}
details.collapsible summary::before {{ content: '▶'; font-size: 10px; color: #5a7a9a; transition: transform 0.2s; }}
details.collapsible[open] summary::before {{ transform: rotate(90deg); }}
details.collapsible[open] summary {{ border-bottom: 1px solid #1a2a4a; padding-bottom: 14px; }}
details.collapsible .section-content {{ padding: 16px 24px; }}
details.collapsible .badge {{ background: #00d4ff; color: #0a0e1a; font-size: 10px; font-weight: 700; padding: 1px 10px; border-radius: 12px; margin-left: auto; }}

/* DNS Records */
.dns-records-table td {{ vertical-align: top; }}
.dh {{ font-family: 'JetBrains Mono', monospace; font-size: 12px; color: #00d4ff; white-space: nowrap; vertical-align: top; }}
.dr-cell {{ padding: 8px 12px; }}
.dr {{ display: flex; align-items: baseline; gap: 8px; padding: 2px 0; font-family: 'JetBrains Mono', monospace; font-size: 12px; }}
.rt {{ color: #5a7a9a; font-weight: 600; min-width: 40px; font-size: 10px; letter-spacing: 1px; }}
.rv {{ color: #c8d6e5; word-break: break-all; }}

/* SSL */
.ssl-card {{ background: rgba(0,0,0,0.2); border: 1px solid #1a2a4a; border-radius: 8px; padding: 14px 16px; margin-bottom: 10px; }}
.ssl-card.expired {{ border-color: rgba(255,68,68,0.3); }}
.ssl-card.valid {{ border-color: rgba(0,255,136,0.2); }}
.ssl-host {{ font-size: 14px; font-weight: 700; color: #00d4ff; margin-bottom: 6px; }}
.ssl-subj {{ font-size: 12px; color: #c8d6e5; margin-bottom: 6px; }}
.ssl-d {{ font-size: 11px; color: #5a7a9a; margin-bottom: 2px; }}
.ssl-d span {{ font-weight: 600; }}
.ssl-d code {{ color: #ffd93d; background: rgba(255,217,61,0.06); padding: 0 4px; border-radius: 2px; }}
.days-valid {{ color: #00ff88; }}
.days-expired {{ color: #ff4444; }}
.sans {{ color: #c8d6e5; font-size: 11px; word-break: break-all; }}

/* Findings table */
.findings-table td {{ font-size: 12px; }}
.findings-table tr:hover td {{ background: rgba(0,212,255,0.03); }}

/* Dir table */
.dirs-table td {{ font-family: 'JetBrains Mono', monospace; font-size: 12px; }}

/* ASN */
.asn-table td {{ padding: 6px 12px; font-size: 13px; }}

.footer {{ text-align: center; padding: 24px; color: #3a5a7a; font-size: 12px; border-top: 1px solid #1a2a4a; }}
.footer strong {{ color: #00d4ff; }}
.endpoint-link {{ color: #00d4ff; text-decoration: none; font-family: 'JetBrains Mono', monospace; font-size: 12px; word-break: break-all; }}
.endpoint-link:hover {{ text-decoration: underline; }}
</style>
</head>
<body>
<div class="header">
<div class="header-content">
<div>
<h1>XREFS0<span class="dot">.</span></h1>
<div class="meta">Target: <strong>{self.domain}</strong> &middot; {self.timestamp} &middot; Author: <strong>MASA</strong></div>
</div>
<div class="badge-group">
{"".join(f'<span class="badge critical">{critical_count} Critical</span>' for _ in range(1) if critical_count > 0)}
{"".join(f'<span class="badge high">{high_count} High</span>' for _ in range(1) if high_count > 0)}
{f'<span class="badge info">{medium_count} Medium</span>' if medium_count > 0 else ''}
</div>
</div>
</div>

<div class="container">

<div class="stats-grid">
<div class="stat-card"><div class="value">{subs_count}</div><div class="label">Subdomains</div></div>
<div class="stat-card"><div class="value">{alive_count}</div><div class="label">Live Hosts</div></div>
<div class="stat-card {'danger' if ports_count > 0 else 'success'}"><div class="value">{ports_count}</div><div class="label">Open Ports</div></div>
<div class="stat-card"><div class="value">{emails_count}</div><div class="label">Emails</div></div>
<div class="stat-card {'danger' if takeover_count > 0 else 'success'}"><div class="value">{takeover_count}</div><div class="label">Takeovers</div></div>
<div class="stat-card"><div class="value">{dir_count}</div><div class="label">Directories</div></div>
<div class="stat-card"><div class="value">{tech_count}</div><div class="label">Tech Stack</div></div>
<div class="stat-card {'danger' if whois_status == 'EXPIRED' else 'success'}"><div class="value">{whois_status}</div><div class="label">Domain Status</div></div>
</div>

<div class="section">
<h2>Subdomains <span class="badge">{subs_count}</span></h2>
<table><tr><th>Hostname</th><th>IP</th><th>Status</th><th>Server</th><th>Title</th></tr>
{sections['subdomains_table_rows']}
</table>
</div>

{sections['real_ip_html']}
{sections['whois_html']}
{sections['ports_html']}
{sections['mail_html']}
{sections['tech_html']}
{sections['cve_html']}
{sections['takeover_html']}
{sections['cloud_html']}
{sections['cms_html']}
{sections['waf_html']}
{sections['cicd_html']}
{sections['api_html']}
{sections['docker_html']}

{sections['dns_records_html']}
{sections['ssl_html']}
{sections['asn_html']}
{sections['emails_html']}
{sections['js_html']}
{sections['wayback_html']}
{sections['commoncrawl_html']}
{sections['dorks_html']}
{sections['github_html']}
{sections['sec_headers_html']}
{sections['cors_html']}
{sections['cookies_html']}
{sections['favicon_html']}
{sections['screenshots_html']}
{sections['spider_html']}
{sections['s3_html']}
{sections['dirs_html']}
{sections['findings_html']}

</div>

<div class="footer">XREFS0 &mdash; Cyber Intelligence &amp; Domain Reconnaissance &mdash; <strong>MASA</strong></div>
</body>
</html>"""
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html)
        console.print(f"[green]HTML report exported to: {filepath}[/green]")

    def export_markdown(self, filepath):
        d = self.data
        lines = []
        lines.append(f"# XREFS0 - {self.domain}")
        lines.append(f"**Generated:** {self.timestamp} | **Author:** MASA")
        lines.append("")
        lines.append("## Summary")
        lines.append(f"- Subdomains: {len(d.get('subdomains', []))}")
        lines.append(f"- Live Hosts: {sum(1 for h in d.get('http', {}).values() if h.get('alive'))}")
        lines.append(f"- Open Ports: {len(d.get('port_scan', {}).get('results', []))}")
        lines.append(f"- Emails: {len(d.get('emails', []))}")
        lines.append(f"- Takeover Vulns: {len(d.get('takeover', []))}")
        whois = d.get("whois", {})
        whois_status = "ACTIVE" if whois.get("is_registered") and not whois.get("is_expired") else "EXPIRED" if whois.get("is_expired") else "UNKNOWN"
        lines.append(f"- Domain Status: {whois_status}")
        lines.append("")
        subdomains = d.get("subdomains", [])
        if subdomains:
            lines.append("## Subdomains")
            lines.append("| Subdomain | IP | Status |")
            lines.append("|-----------|----|--------|")
            for s in subdomains[:50]:
                ip = ""
                status = ""
                if s in d.get("dns", {}):
                    ips = d["dns"][s].get("A", [])
                    if ips:
                        ip = ips[0]
                if s in d.get("http", {}):
                    h = d["http"][s]
                    status = str(h.get("status_code", ""))
                lines.append(f"| {s} | {ip} | {status} |")
            lines.append("")
        ports = d.get("port_scan", {}).get("results", [])
        if ports:
            lines.append("## Open Ports")
            lines.append("| Port | Service | Banner |")
            lines.append("|------|---------|--------|")
            for p in ports[:30]:
                lines.append(f"| {p['port']} | {p.get('service', '')} | {p.get('banner', '')[:50]} |")
            lines.append("")
        emails = d.get("emails", [])
        if emails:
            lines.append("## Emails Found")
            for e in emails:
                lines.append(f"- {e}")
            lines.append("")
        lines.append("---")
        lines.append(f"*XREFS0 - Cyber Intelligence & Domain Recon Platform | Generated by MASA*")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        console.print(f"[green]Markdown exported to: {filepath}[/green]")

    def export_pdf(self, filepath):
        try:
            import subprocess
            html_path = filepath.replace(".pdf", ".html")
            self.export_html(html_path)
            result = subprocess.run(
                ["node", "-e", f"""
const {{ promises: fs }} = require('fs');
(async () => {{
  try {{
    const puppeteer = require('puppeteer');
    const browser = await puppeteer.launch({{ headless: 'new' }});
    const page = await browser.newPage();
    const html = fs.readFile('{html_path.replace(os.sep, '/')}', 'utf8');
    await page.setContent(html);
    await page.pdf({{ path: '{filepath.replace(os.sep, '/')}', format: 'A4' }});
    await browser.close();
  }} catch(e) {{}}
}})();
"""],
                capture_output=True,
                timeout=30,
            )
            if os.path.exists(filepath) and os.path.getsize(filepath) > 0:
                console.print(f"[green]PDF exported to: {filepath}[/green]")
            else:
                console.print("[yellow]PDF export requires puppeteer. Falling back to HTML only.[/yellow]")
        except Exception:
            self.export_html(filepath.replace(".pdf", ".html"))
            console.print("[yellow]PDF export requires puppeteer. HTML exported instead.[/yellow]")
